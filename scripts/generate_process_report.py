import os
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

if sys.stdout.encoding != 'utf-8':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

def generate_report():
    wb = openpyxl.Workbook()
    
    # ==========================================
    # SHEET 1: PROCESS REPORT (CHUẨN FPT EXCEL)
    # ==========================================
    ws = wb.active
    ws.title = '1. Báo Cáo Tiến Độ'
    ws.views.sheetView[0].showGridLines = True

    # Styles & Colors
    navy_header_fill = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')
    blue_sub_fill = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')
    zebra_fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')

    white_bold_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    title_font = Font(name='Calibri', size=15, bold=True, color='1F497D')
    subtitle_font = Font(name='Calibri', size=10, bold=True, color='366092')
    bold_font = Font(name='Calibri', size=10, bold=True)
    regular_font = Font(name='Calibri', size=10)

    thin_border = Border(
        left=Side(style='thin', color='B0C4DE'),
        right=Side(style='thin', color='B0C4DE'),
        top=Side(style='thin', color='B0C4DE'),
        bottom=Side(style='thin', color='B0C4DE')
    )

    # Title Block
    ws.merge_cells('A1:U1')
    ws['A1'] = 'BÁO CÁO TIẾN ĐỘ THỰC HIỆN ĐỒ ÁN TỐT NGHIỆP (CAPSTONE PROCESS REPORT)'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:U2')
    ws['A2'] = 'Code: IAP491_FA26_PI_GUARD | Học kỳ Fall 2026 (07/09/2026 – 20/12/2026) | Khoa An toàn Thông tin - Đại học FPT'
    ws['A2'].font = subtitle_font
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A3:U3')
    ws['A3'] = 'Topic: A Machine-Learning Guardrail for Detecting Prompt Injection and Jailbreak Attacks on LLM Applications (PI-Guard)'
    ws['A3'].font = Font(name='Calibri', size=10, italic=True, bold=True, color='1F497D')
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')

    # Headers Row 4 & 5
    ws.merge_cells('A4:A5')
    ws['A4'] = 'MSSV'
    ws.merge_cells('B4:B5')
    ws['B4'] = 'Họ và tên'

    ws.merge_cells('C4:F4')
    ws['C4'] = 'Phân công nhiệm vụ cốt lõi'
    ws['C5'] = 'Nhiệm vụ 1'
    ws['D5'] = 'Nhiệm vụ 2'
    ws['E5'] = 'Nhiệm vụ 3'
    ws['F5'] = 'Nhiệm vụ 4'

    ws.merge_cells('G4:U4')
    ws['G4'] = 'Tiến độ thực hiện hàng tuần (Tuần 1: 07/09/2026 – Tuần 15: 20/12/2026)'

    weeks = [
        ('Tuần 1', '07/09 - 13/09'),
        ('Tuần 2', '14/09 - 20/09'),
        ('Tuần 3', '21/09 - 27/09'),
        ('Tuần 4', '28/09 - 04/10'),
        ('Tuần 5', '05/10 - 11/10'),
        ('Tuần 6', '12/10 - 18/10'),
        ('Tuần 7', '19/10 - 25/10'),
        ('Tuần 8', '26/10 - 01/11'),
        ('Tuần 9', '02/11 - 08/11'),
        ('Tuần 10', '09/11 - 15/11'),
        ('Tuần 11', '16/11 - 22/11'),
        ('Tuần 12', '23/11 - 29/11'),
        ('Tuần 13', '30/11 - 06/12'),
        ('Tuần 14', '07/12 - 13/12'),
        ('Tuần 15', '14/12 - 20/12')
    ]
    
    for idx, (w_name, w_dates) in enumerate(weeks):
        col = get_column_letter(7 + idx)
        ws[f'{col}5'] = f"{w_name}\n({w_dates})"

    for col_idx in range(1, 22):
        c1 = ws.cell(row=4, column=col_idx)
        c2 = ws.cell(row=5, column=col_idx)
        c1.fill = navy_header_fill
        c1.font = white_bold_font
        c1.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c2.fill = blue_sub_fill
        c2.font = bold_font
        c2.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c1.border = thin_border
        c2.border = thin_border

    # Member data
    members = [
        {
            'mssv': 'SE182034',
            'name': 'Nguyễn Văn Trường (Leader)',
            'nv1': 'Thiết kế kiến trúc hệ thống PI-Guard, thu thập và xây dựng Dataset đa nguồn (Hugging Face).',
            'nv2': 'Nghiên cứu & triển khai thuật toán Group-Aware Splitting chống rò rỉ dữ liệu (Data Leakage).',
            'nv3': 'Phụ trách Report No.1 (Introduction) & Report No.2 (Literature Review).',
            'nv4': 'Điều phối tiến độ nhóm, chuẩn hóa hồ sơ Review 1 - 4 và nộp báo cáo tuần cho GVHD.',
            'w1': 'Khởi động & Sàng lọc Papers:\n- Họp GVHD Meeting 1 (29/08) & Meeting 2 (01/09).\n- Đọc & sàng lọc 10 papers theo Register.\n- Cập nhật cơ sở lý thuyết Chapter 1&2.',
            'w2': 'Bối cảnh & Phân loại:\n- Soạn thảo Chapter 1 (Introduction).\n- Phân loại Threat Taxonomy (OWASP).\n- Lỗ hổng Von Neumann NLP.',
            'w3': 'Literature Review:\n- Soạn thảo Chapter 2 (Literature Review).\n- Khảo sát SOTA Guardrails.\n- Dàn ý slide 9 trang Review 1.',
            'w4': 'BẢO VỆ REVIEW 1 (GVHD):\n- Hoàn thiện Report No.1 & No.2.\n- Thiết kế Slide PPTX & tập diễn tập.\n- Bảo vệ thành công Review 1.',
            'w5': 'Data Engineering:\n- Tải 5 datasets từ Hugging Face.\n- Cài đặt Group-Aware Split.\n- Làm sạch & khử trùng lặp dữ liệu.',
            'w6': 'Hỗ trợ Baseline ML:\n- Phân tích phân phối nhãn dữ liệu.\n- Đo lường độ rò rỉ Inter-cluster.\n- Hỗ trợ xây dựng baseline TF-IDF.',
            'w7': 'Soạn thảo & Cập nhật Docs:\n- Soạn thảo Chapter 3 (Methodology).\n- Tổng hợp cấu trúc Report No.3.\n- Chuẩn bị tài liệu & slide Review 2.',
            'w8': 'BẢO VỆ REVIEW 2 (GVHD):\n- Báo cáo Chapter 3 (Methodology & ML).\n- Tiếp thu góp ý & cập nhật hoàn thiện docs.\n- Nghiệm thu Report No.3 (20% điểm).',
            'w9': 'Transformer Training:\n- Hỗ trợ fine-tuning DeBERTa-v3.\n- Phân tích Attention Weight.\n- Đánh giá độ bền Adversarial.',
            'w10': 'Hoàn thiện Report No.4:\n- Đo lường chỉ số F1, FPR trên CPU.\n- Kiểm toán kết quả nén ONNX INT8.\n- Nộp Report No.4 cho GVHD.',
            'w11': 'Tích hợp API Middleware:\n- Điều phối tích hợp FastAPI proxy.\n- Đo latency end-to-end P95 < 30ms.\n- Dựng dashboard Streamlit.',
            'w12': 'Chuẩn bị Hội Đồng 1:\n- Hoàn thiện tài liệu kỹ thuật demo.\n- Tổng duyệt Prototype & Slide Hội đồng.\n- Kiểm tra độ ổn định hệ thống.',
            'w13': 'BÁO CÁO HỘI ĐỒNG 1 (MIDTERM):\n- Bảo vệ Prototype Demo trước Hội đồng.\n- Báo cáo kết quả thực nghiệm Chapter 4.\n- Tiếp thu ý kiến phản biện Hội đồng.',
            'w14': 'Discussion, Conclusion & Thesis:\n- Soạn thảo Report No.5 & No.6.\n- Tổng hợp Master Thesis 6 chương.\n- Quét Turnitin (< 20%), duyệt Slide.',
            'w15': 'BÁO CÁO HỘI ĐỒNG FINAL (TỐT NGHIỆP):\n- Tổng duyệt Slide PPTX bảo vệ.\n- BẢO VỆ CHÍNH THỨC TRƯỚC HỘI ĐỒNG.\n- Hoàn tất hồ sơ tốt nghiệp FPT.'
        },
        {
            'mssv': 'SE182087',
            'name': 'Nguyễn Quí Đức',
            'nv1': 'Nghiên cứu lý thuyết Classical Machine Learning và biểu diễn đặc trưng ngôn ngữ.',
            'nv2': 'Xây dựng pipeline trích xuất đặc trưng kết hợp (Word + Char n-grams TF-IDF).',
            'nv3': 'Huấn luyện và đánh giá các mô hình Baseline (Logistic Regression, LinearSVC, XGBoost).',
            'nv4': 'Phụ trách Report No.3 (Methodology - Baseline ML & Feature Engineering).',
            'w1': 'Khảo sát ML & JailGuard:\n- Tham gia Meeting 1 & Meeting 2 (01/09).\n- Đọc & phân tích JailGuard (TOSEM 2025).\n- Tìm hiểu Targeted Mutators Workflow.',
            'w2': 'Threat Model & Attack Surface:\n- Xây dựng sơ đồ NIST AI 100-2e2025.\n- Khóa chặt Attack Surface /v1/chat.\n- Thiết kế kiến trúc bảo vệ 3 lớp.',
            'w3': 'Chuẩn bị Review 1:\n- Chuẩn bị slide Threat Model & 3L.\n- Xây dựng script baseline TF-IDF mẫu.\n- Thử nghiệm feature extraction.',
            'w4': 'BẢO VỆ REVIEW 1 (GVHD):\n- Trình bày Threat Model & Kiến trúc 3L.\n- Phản biện phương pháp luận baseline.\n- Hoàn thành mốc Review 1.',
            'w5': 'Trích xuất đặc trưng:\n- Xây dựng Word+Char TF-IDF pipeline.\n- Tối ưu hóa sublinear_tf, max_features.\n- Xử lý bộ tiền lọc Base64/Leet.',
            'w6': 'Huấn luyện Baseline ML:\n- Huấn luyện LR, LinearSVC, XGBoost.\n- Đánh giá F1, FPR, độ trễ CPU.\n- So sánh hiệu năng các mô hình.',
            'w7': 'Methodology & Baseline ML:\n- Soạn thảo chi tiết Chapter 3.\n- Tổng hợp số liệu F1/FPR Baseline.\n- Cập nhật tài liệu kỹ thuật TF-IDF.',
            'w8': 'BẢO VỆ REVIEW 2 (GVHD):\n- Thuyết trình phương pháp luận Chapter 3.\n- Demo kết quả Baseline ML (LR, SVC, XGB).\n- Cập nhật docs theo góp ý của Thầy.',
            'w9': 'Phân tích Robustness:\n- Đánh giá khả năng chống nhiễu ký tự.\n- Kiểm thử bộ lọc Base64 decoder.\n- Hỗ trợ đánh giá DeBERTa-v3.',
            'w10': 'Báo cáo No.4:\n- Tổng hợp số liệu đối chuẩn Baseline.\n- Vẽ biểu đồ ROC-AUC, Confusion Matrix.\n- Đóng góp phần Methodology Report 4.',
            'w11': 'Tích hợp Model vào API:\n- Đóng gói mô hình joblib sang API.\n- Viết routing xử lý phân loại nhanh.\n- Kiểm thử latency tầng baseline.',
            'w12': 'Chuẩn bị Hội Đồng 1:\n- Đóng gói tài liệu trích xuất đặc trưng.\n- Kiểm thử chéo mô hình Baseline vs ONNX.\n- Tham gia tổng duyệt demo.',
            'w13': 'BÁO CÁO HỘI ĐỒNG 1 (MIDTERM):\n- Trình bày kiến trúc ML & kết quả.\n- Phản biện thuật toán trích xuất đặc trưng.\n- Tiếp thu ý kiến Hội đồng.',
            'w14': 'Soạn thảo Luận văn:\n- Soạn thảo chuyên sâu Chapter 3 & 4.\n- So sánh ưu nhược TF-IDF vs Transformer.\n- Viết phần thảo luận giới hạn mô hình.',
            'w15': 'BÁO CÁO HỘI ĐỒNG FINAL (TỐT NGHIỆP):\n- Phụ trách báo cáo phần Methodology.\n- Trả lời chất vấn của Hội đồng.\n- Bảo vệ tốt nghiệp xuất sắc.'
        },
        {
            'mssv': 'SE181851',
            'name': 'Phạm Minh Hoàng Việt',
            'nv1': 'Nghiên cứu cơ chế Disentangled Attention của mô hình microsoft/deberta-v3-base.',
            'nv2': 'Thực hiện Supervised Fine-Tuning Transformer và tối ưu hàm mất mát BCEWithLogitsLoss.',
            'nv3': 'Lượng hóa động ONNX INT8 Runtime để tối ưu độ trễ P95 < 30ms trên CPU.',
            'nv4': 'Phụ trách Report No.4 (Experimental and Results - Training & Adversarial Tests).',
            'w1': 'Khảo sát RAP-ID & BIPIA:\n- Tham gia Meeting 1 & Meeting 2 (01/09).\n- Phân tích Pre-fill pass (RAP-ID) & BIPIA.\n- Tìm cơ sở khoa học DeBERTa < 600M.',
            'w2': 'SOTA Matrix & Target LLM:\n- Xây dựng Model Selection Matrix.\n- Thiết lập benchmark 5 Target LLMs.\n- Soạn thảo phân tích đối chuẩn.',
            'w3': 'Demo Scenarios Review 1:\n- Thiết kế ma trận 4 Kịch bản Demo 2x2.\n- Viết kịch bản tấn công Prompt Injection.\n- Chuẩn bị slide trình bày SOTA.',
            'w4': 'BẢO VỆ REVIEW 1 (GVHD):\n- Trình bày Model Selection & Demo 2x2.\n- Bảo vệ lựa chọn DeBERTa-v3.\n- Hoàn thành mốc Review 1.',
            'w5': 'Chuẩn bị Fine-Tuning:\n- Chuẩn bị dataloader cho Transformer.\n- Thiết lập pipeline tokenize độ dài 512.\n- Cấu hình hyperparameters AdamW/Cosine.',
            'w6': 'Khảo sát Baseline & Transformer:\n- Hỗ trợ đánh giá Baseline ML.\n- Khảo sát Disentangled Attention.\n- Chuẩn bị cấu trúc Chapter 3.',
            'w7': 'Hỗ trợ Thực nghiệm & Docs:\n- Soạn thảo cấu trúc Chapter 3.\n- Chuẩn bị ma trận đánh giá đối chuẩn.\n- Kiểm tra tính nhất quán dữ liệu.',
            'w8': 'BẢO VỆ REVIEW 2 & Khởi động SFT:\n- Báo cáo định hướng DeBERTa-v3.\n- Cập nhật docs Chapter 3 sau Review 2.\n- Thiết lập pipeline fine-tuning GPU.',
            'w9': 'Fine-Tuning DeBERTa-v3:\n- Huấn luyện DeBERTa-v3 trên GPU.\n- Tối ưu BCEWithLogitsLoss chống lệch nhãn.\n- Theo dõi validation loss & F1-score.',
            'w10': 'Lượng hóa ONNX INT8:\n- Xuất mô hình sang ONNX Runtime.\n- Áp dụng dynamic INT8 quantization.\n- Đạt độ trễ P95 < 30ms trên CPU. Nộp R4.',
            'w11': 'Tích hợp Engine ONNX:\n- Viết wrapper suy luận ONNX cho API.\n- Tối ưu hóa đa luồng CPU (OMP_NUM_THREADS).\n- Đo đạc thông lượng xử lý RPS.',
            'w12': 'Chuẩn bị Hội Đồng 1:\n- Kiểm thử hiệu năng ONNX trên máy trạm.\n- Chuẩn bị slide thực nghiệm Chapter 4.\n- Tổng duyệt phần trình bày AI Model.',
            'w13': 'BÁO CÁO HỘI ĐỒNG 1 (MIDTERM):\n- Trình diễn trực tiếp mô hình ONNX INT8.\n- Chứng minh độ trễ P95 < 30ms trên CPU.\n- Bảo vệ thành công Prototype.',
            'w14': 'Biên soạn Kết quả Thực nghiệm:\n- Viết chi tiết toàn văn Chapter 4.\n- Bảng đối chuẩn F1, FPR, Latency.\n- Phân tích trade-off độ chính xác - tốc độ.',
            'w15': 'BÁO CÁO HỘI ĐỒNG FINAL (TỐT NGHIỆP):\n- Trình bày kiến trúc DeBERTa & ONNX INT8.\n- Trả lời chất vấn chuyên sâu về AI Security.\n- Hoàn thành xuất sắc đồ án.'
        },
        {
            'mssv': 'SE180235',
            'name': 'Đỗ Đoàn Duy Phương',
            'nv1': 'Phát triển Asynchronous API Middleware bằng FastAPI tích hợp chốt chặn Guardrail.',
            'nv2': 'Xây dựng giao diện Dashboard tương tác trực quan bằng Streamlit phục vụ kiểm thử.',
            'nv3': 'Định dạng và biên tập toàn văn Luận văn tốt nghiệp theo chuẩn FPT IAP491.',
            'nv4': 'Phụ trách Report No.5 (Discussion) & Report No.6 (Conclusion & Slide Deck).',
            'w1': 'Khảo sát Jailbreak & Do-Not-Answer:\n- Tham gia Meeting 1 & Meeting 2 (01/09).\n- Phân loại 3 chiến thuật Jailbreak (Survey).\n- Khảo sát Do-Not-Answer & Black-box def.',
            'w2': 'Research Questions & Gaps:\n- Soạn thảo 3 Câu hỏi IEEE (RQ1-RQ3).\n- Xác định 3 Gaps & 4 Đóng góp mới.\n- Định vị ranh giới In-scope/Out-of-scope.',
            'w3': 'Slide Outline Review 1:\n- Thiết kế bộ slide 9 trang Review 1.\n- Dựng demo UI mô phỏng 4 kịch bản.\n- Soạn kịch bản thuyết trình 15 phút.',
            'w4': 'BẢO VỆ REVIEW 1 (GVHD):\n- Trình bày RQ1-RQ3 & Đóng góp mới.\n- Điều phối trình chiếu slide Review 1.\n- Hoàn thành mốc Review 1.',
            'w5': 'Xây dựng API Core:\n- Viết endpoint POST /v1/chat/completions.\n- Xây dựng tầng tiền xử lý regex & decoding.\n- Tích hợp cấu trúc phản hồi chuẩn RFC.',
            'w6': 'Tích hợp Baseline vào API:\n- Kết nối mô hình TF-IDF vào middleware.\n- Ghi log request/response vào JSONL.\n- Xử lý bất đồng bộ async/await.',
            'w7': 'Kiến trúc Middleware & Docs:\n- Soạn thảo kiến trúc API Chapter 3.\n- Thiết kế bộ slide báo cáo Review 2.\n- Hoàn thiện bản thảo Report No.3.',
            'w8': 'BẢO VỆ REVIEW 2 (GVHD):\n- Trình diễn kết nối API với Baseline ML.\n- Cập nhật docs theo kết luận họp.\n- Bắt đầu dựng Streamlit Dashboard.',
            'w9': 'Phát triển Streamlit Dashboard:\n- Dựng UI nhập prompt kiểm thử trực tiếp.\n- Hiển thị nhãn, confidence score, latency.\n- Tạo trang so sánh 4 kịch bản demo.',
            'w10': 'Tích hợp ONNX Engine:\n- Kết nối mô hình ONNX INT8 vào API.\n- Thêm tính năng toggle chuyển đổi mô hình.\n- Tối ưu hóa xử lý lỗi HTTP 400/500.',
            'w11': 'Hoàn thiện Dashboard:\n- Thêm biểu đồ phân tích và giám sát log.\n- Dựng dashboard đo latency & throughput.\n- Đóng gói Docker Compose tiện lợi.',
            'w12': 'Chuẩn bị Hội Đồng 1:\n- Hoàn thiện UI Dashboard tương tác cao.\n- Viết kịch bản live demo 4 kịch bản.\n- Tổng duyệt kỹ thuật cùng nhóm.',
            'w13': 'BÁO CÁO HỘI ĐỒNG 1 (MIDTERM):\n- Trình chiếu trực tiếp Dashboard & API.\n- Demo tương tác trực quan 4 kịch bản.\n- Tiếp thu góp ý từ Hội đồng Giữa kỳ.',
            'w14': 'Soạn thảo Report No.5 & No.6:\n- Viết Chapter 5 (Discussion & Limitations).\n- Viết Chapter 6 (Conclusion & Future Work).\n- Định dạng chuẩn IEEE/FPT cho Luận văn.',
            'w15': 'BÁO CÁO HỘI ĐỒNG FINAL (TỐT NGHIỆP):\n- Trình bày Demo & Kiến trúc phần mềm.\n- Điều phối phần trình chiếu Hội đồng.\n- Đạt kết quả Tốt nghiệp xuất sắc.'
        }
    ]

    for row_idx, mem in enumerate(members, start=6):
        ws.cell(row=row_idx, column=1, value=mem['mssv'])
        ws.cell(row=row_idx, column=2, value=mem['name'])
        ws.cell(row=row_idx, column=3, value=mem['nv1'])
        ws.cell(row=row_idx, column=4, value=mem['nv2'])
        ws.cell(row=row_idx, column=5, value=mem['nv3'])
        ws.cell(row=row_idx, column=6, value=mem['nv4'])
        
        for w_idx in range(1, 16):
            w_key = f'w{w_idx}'
            ws.cell(row=row_idx, column=6 + w_idx, value=mem.get(w_key, '[ ] Kế hoạch theo lộ trình'))

        for c in range(1, 22):
            cell = ws.cell(row=row_idx, column=c)
            cell.font = regular_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            if row_idx % 2 == 1:
                cell.fill = zebra_fill

    # ==========================================
    # SHEET 2: INTERACTIVE CHECKLIST (23 TASKS)
    # ==========================================
    ws_check = wb.create_sheet(title='2. Checklist Tiến Độ')
    ws_check.views.sheetView[0].showGridLines = True

    ws_check.merge_cells('A1:G1')
    ws_check['A1'] = 'DANH MỤC CÔNG VIỆC CHI TIẾT & CHECKLIST TIẾN ĐỘ (PI-GUARD CAPSTONE)'
    ws_check['A1'].font = title_font
    ws_check['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws_check.merge_cells('A2:G2')
    ws_check['A2'] = 'Học kỳ Fall 2026 (07/09/2026 – 20/12/2026) | Hướng dẫn: Chọn trạng thái tại cột D từ danh sách thả xuống'
    ws_check['A2'].font = subtitle_font
    ws_check['A2'].alignment = Alignment(horizontal='center', vertical='center')

    check_headers = ['Mã Task', 'Tuần / Cột mốc', 'Nội dung công việc', 'Trạng thái', 'Thành viên phụ trách', 'Sản phẩm đầu ra', 'Ghi chú / Deadline']
    for col_idx, h in enumerate(check_headers, start=1):
        c = ws_check.cell(row=4, column=col_idx, value=h)
        c.fill = navy_header_fill
        c.font = white_bold_font
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = thin_border

    # Data validation dropdown for Status
    dv = DataValidation(type='list', formula1='"Hoàn thành,Đang thực hiện,Chưa bắt đầu"', allow_blank=True)
    ws_check.add_data_validation(dv)
    dv.add('D5:D50')

    tasks = [
        ('T01', 'Tuần 1 (07/09 - 13/09)', 'Họp GVHD định hướng bài toán (Meeting 1_29_08) & Sàng lọc 10 papers khoa học (Meeting 2_01_09)', 'Hoàn thành', 'Cả 4 thành viên', 'Biên bản Meeting 1 & Meeting 2.md', '13/09/2026'),
        ('T02', 'Tuần 1 (07/09 - 13/09)', 'Thu thập & thẩm định 17 papers chuẩn IEEE >= 2022 theo CAPSTONE REGISTER', 'Hoàn thành', 'Trường (Leader)', 'References/ & REFERENCES_LOG.md', '13/09/2026'),
        ('T03', 'Tuần 2 (14/09 - 20/09)', 'Soạn thảo Chapter 1: Background & Problem Statement (Lỗ hổng Von Neumann NLP)', 'Đang thực hiện', 'Trường', 'workspaces/truongnv/docs/', '16/09/2026'),
        ('T04', 'Tuần 2 (14/09 - 20/09)', 'Phân loại Threat Taxonomy (Direct/Indirect Injection vs Jailbreak theo OWASP)', 'Đang thực hiện', 'Trường & Đức', 'workspaces/truongnv/docs/', '17/09/2026'),
        ('T05', 'Tuần 2 (14/09 - 20/09)', 'Xây dựng Threat Model (NIST AI 100-2e2025) & Attack Surface (/v1/chat)', 'Đang thực hiện', 'Đức', 'workspaces/ducnq/', '18/09/2026'),
        ('T06', 'Tuần 2 (14/09 - 20/09)', 'Thiết kế Kiến trúc bảo vệ 3 lớp & Cơ chế phòng thủ độ bền Robustness', 'Đang thực hiện', 'Đức', 'workspaces/ducnq/', '19/09/2026'),
        ('T07', 'Tuần 2 (14/09 - 20/09)', 'Soạn thảo 3 Câu hỏi nghiên cứu IEEE (RQ1-RQ3), 3 Gaps & 4 Đóng góp mới', 'Đang thực hiện', 'Phương', 'workspaces/phuongddd/', '20/09/2026'),
        ('T08', 'Tuần 3 (21/09 - 27/09)', 'Khảo sát SOTA Guardrails, Model Selection Matrix & Benchmark 5 Target LLMs', 'Đang thực hiện', 'Việt', 'workspaces/vietpmh/', '23/09/2026'),
        ('T09', 'Tuần 3 (21/09 - 27/09)', 'Thiết kế & kiểm thử Ma trận 4 Kịch bản Demo (2x2: Vulnerable vs Protected)', 'Đang thực hiện', 'Việt & Phương', 'workspaces/vietpmh/', '25/09/2026'),
        ('T10', 'Tuần 3 (21/09 - 27/09)', 'Hoàn thiện toàn văn Report No.1: Introduction (Chapter 1 — 10% Process Mark)', 'Đang thực hiện', 'Trường (Leader)', 'workspaces/truongnv/docs/', '26/09/2026'),
        ('T11', 'Tuần 3 (21/09 - 27/09)', 'Hoàn thiện toàn văn Report No.2: Literature Review (Chapter 2 — 25% Process Mark)', 'Đang thực hiện', 'Trường & Phương', 'workspaces/truongnv/docs/', '27/09/2026'),
        ('T12', 'Tuần 3 (21/09 - 27/09)', 'Thiết kế dàn ý slide 9 trang Review 1 Presentation Slides (Bao gồm 2 Chương)', 'Đang thực hiện', 'Phương', 'workspaces/truongnv/docs/', '27/09/2026'),
        ('T13', 'Tuần 4 (28/09 - 04/10)', 'Thiết kế slide PowerPoint (.pptx) & Tập dượt thuyết trình 15 phút (2 Chương)', 'Chưa bắt đầu', 'Cả 4 thành viên', 'Slide PPTX Review 1', '02/10/2026'),
        ('T14', 'Tuần 4 (28/09 - 04/10)', 'Họp tổng kết tuần, cập nhật Process Report Excel & Nộp Report 1 & 2 cho GVHD', 'Chưa bắt đầu', 'Trường (Leader)', 'PI_GUARD_PROCESS_REPORT.xlsx', '03/10/2026'),
        ('T15', 'Tuần 4 (28/09 - 04/10)', 'CỘT MỐC 1 — BẢO VỆ REVIEW 1 TRƯỚC GVHD (CHAPTERS 1 & 2)', 'Chưa bắt đầu', 'Cả 4 thành viên', 'Biên bản nghiệm thu Review 1', '04/10/2026'),
        ('T16', 'Tuần 5 - 6 (05/10 - 18/10)', 'Thu thập Dataset đa nguồn, Group-Aware Split & Huấn luyện Baseline ML', 'Chưa bắt đầu', 'Trường & Đức', 'data/processed/ & models/baseline/', '18/10/2026'),
        ('T17', 'Tuần 7 (19/10 - 25/10)', 'Soạn thảo, cập nhật docs Chapter 3 (Methodology - Report No.3) & Chuẩn bị Slide Review 2', 'Chưa bắt đầu', 'Đức & Trường', 'docs/thesis/chapters/03_Methodology.md & Slide Review 2', '25/10/2026'),
        ('T18', 'Tuần 8 (26/10 - 01/11)', 'CỘT MỐC 2 — BẢO VỆ REVIEW 2 TRƯỚC GVHD (CHAPTER 3), Nộp Report No.3 & Cập nhật Docs hoàn chỉnh', 'Chưa bắt đầu', 'Cả 4 thành viên', 'Biên bản nghiệm thu Review 2 & Chapter 3 final', '01/11/2026'),
        ('T19', 'Tuần 9 - 10 (02/11 - 15/11)', 'Fine-tuning DeBERTa-v3, Lượng hóa ONNX INT8, Đánh giá Adversarial & Nộp Report No.4', 'Chưa bắt đầu', 'Việt & Đức', 'Report No.4 & models/onnx/', '15/11/2026'),
        ('T20', 'Tuần 11 - 12 (16/11 - 29/11)', 'Xây dựng API Middleware, Streamlit Dashboard & Tích hợp Prototype Demo 3 lớp', 'Chưa bắt đầu', 'Phương & Việt', 'Hệ thống Prototype hoàn chỉnh', '29/11/2026'),
        ('T21', 'Tuần 13 (30/11 - 06/12)', 'CỘT MỐC 3 — BÁO CÁO HỘI ĐỒNG 1 (HỘI ĐỒNG GIỮA KỲ — PROTOTYPE DEMO & CHAPTER 4)', 'Chưa bắt đầu', 'Cả 4 thành viên', 'Biên bản nghiệm thu Hội đồng 1', '06/12/2026'),
        ('T22', 'Tuần 14 (07/12 - 13/12)', 'Soạn Report No.5 & No.6, Tổng hợp Master Thesis 6 chương qua script & Quét Turnitin (< 20%)', 'Chưa bắt đầu', 'Cả 4 thành viên', 'FINAL_THESIS.md & Report 5, 6', '13/12/2026'),
        ('T23', 'Tuần 15 (14/12 - 20/12)', 'CỘT MỐC 4 — BẢO VỆ TỐT NGHIỆP TOÀN DIỆN 6 CHƯƠNG TRƯỚC HỘI ĐỒNG FINAL', 'Chưa bắt đầu', 'Cả 4 thành viên', 'Final Thesis & Slide Defense', '20/12/2026')
    ]

    for row_idx, task in enumerate(tasks, start=5):
        for col_idx, val in enumerate(task, start=1):
            c = ws_check.cell(row=row_idx, column=col_idx, value=val)
            c.font = regular_font
            c.border = thin_border
            if col_idx == 4:
                c.alignment = Alignment(horizontal='center', vertical='center')
                if val == 'Hoàn thành':
                    c.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
                    c.font = Font(name='Calibri', size=10, bold=True, color='375623')
                elif val == 'Đang thực hiện':
                    c.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
                    c.font = Font(name='Calibri', size=10, bold=True, color='7F6000')
                else:
                    c.fill = zebra_fill
            elif col_idx in [1, 2, 7]:
                c.alignment = Alignment(horizontal='center', vertical='center')
            else:
                c.alignment = Alignment(vertical='center', wrap_text=True)

    # Column dimensions
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 30
    for col_idx in range(7, 22):
        ws.column_dimensions[get_column_letter(col_idx)].width = 36

    ws_check.column_dimensions['A'].width = 10
    ws_check.column_dimensions['B'].width = 26
    ws_check.column_dimensions['C'].width = 48
    ws_check.column_dimensions['D'].width = 20
    ws_check.column_dimensions['E'].width = 18
    ws_check.column_dimensions['F'].width = 32
    ws_check.column_dimensions['G'].width = 16

    # Row heights
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 25
    ws.row_dimensions[5].height = 30
    for r in range(6, 10):
        ws.row_dimensions[r].height = 110

    ws_check.row_dimensions[1].height = 25
    ws_check.row_dimensions[2].height = 20
    ws_check.row_dimensions[4].height = 25
    for r in range(5, len(tasks) + 5):
        ws_check.row_dimensions[r].height = 25

    # Save outputs to all relevant locations
    workspace_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    out_paths = [
        os.path.join(workspace_root, 'reports', 'PI_GUARD_PROCESS_REPORT.xlsx'),
        os.path.join(workspace_root, 'Meeting', 'PI_GUARD_PROCESS_REPORT.xlsx'),
        os.path.join(workspace_root, 'workspaces', 'truongnv', 'Meeting', 'PI_GUARD_PROCESS_REPORT.xlsx')
    ]
    for p in out_paths:
        os.makedirs(os.path.dirname(p), exist_ok=True)
        wb.save(p)
        print(f'Saved: {p}')

if __name__ == '__main__':
    generate_report()
