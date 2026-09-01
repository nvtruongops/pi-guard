import os
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

if sys.stdout.encoding != 'utf-8':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

def create_personal_process_report():
    output_path = r"D:\Work\Do-an\workspaces\truong_data_eng\Meeting\PI_GUARD_PROCESS_REPORT.xlsx"
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb = openpyxl.Workbook()
    
    # ----------------------------------------------------
    # Styles & Colors (Clean Academic / Corporate Theme)
    # ----------------------------------------------------
    font_family = "Segoe UI"
    
    title_font = Font(name=font_family, size=15, bold=True, color="FFFFFF")
    subtitle_font = Font(name=font_family, size=10, italic=True, color="E5E7EB")
    header_font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    section_font = Font(name=font_family, size=11, bold=True, color="1E3A8A")
    regular_font = Font(name=font_family, size=10)
    bold_font = Font(name=font_family, size=10, bold=True)
    
    primary_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Deep Navy
    sub_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")     # Soft Slate Blue
    gray_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")    # Soft Gray
    green_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")   # Soft Green
    yellow_fill = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")  # Soft Yellow
    
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # ====================================================
    # SHEET 1: HỒ SƠ CÁ NHÂN & TIẾN ĐỘ
    # ====================================================
    ws1 = wb.active
    ws1.title = "1. Hồ Sơ Cá Nhân"
    ws1.views.sheetView[0].showGridLines = True
    
    # Title Banner
    ws1.merge_cells("A1:G1")
    ws1["A1"] = "BÁO CÁO TIẾN ĐỘ CÁ NHÂN — NGUYỄN VĂN TRƯỜNG (LEADER)"
    ws1["A1"].font = title_font
    ws1["A1"].fill = primary_fill
    ws1["A1"].alignment = center_align
    ws1.row_dimensions[1].height = 36

    ws1.merge_cells("A2:G2")
    ws1["A2"] = "Khóa luận Tốt nghiệp: PI-Guard | Học kỳ Fall 2026 (07/09/2026 – 20/12/2026) | Khoa An toàn Thông tin — Đại học FPT"
    ws1["A2"].font = subtitle_font
    ws1["A2"].fill = primary_fill
    ws1["A2"].alignment = center_align
    ws1.row_dimensions[2].height = 20

    # Section 1: Thông tin sinh viên
    ws1.merge_cells("A4:G4")
    ws1["A4"] = "I. THÔNG TIN SINH VIÊN & TRÁCH NHIỆM CHÍNH"
    ws1["A4"].font = section_font
    ws1["A4"].fill = sub_fill
    ws1["A4"].alignment = left_align
    ws1.row_dimensions[4].height = 24

    student_info = [
        ("Họ và Tên:", "Nguyễn Văn Trường", "Mã Sinh Viên:", "SE182034", "Vai Trò:", "Trưởng Nhóm (Leader)"),
        ("Chuyên Ngành:", "An Toàn Thông Tin (Information Assurance - IA)", "Khóa / Kỳ:", "K18 / Fall 2026", "Mã Đồ Án:", "IAP491"),
        ("Tên Đề Tài:", "A Machine-Learning Guardrail for Detecting Prompt Injection and Jailbreak Attacks on LLM Applications (PI-Guard)", "", "", "", ""),
        ("Giảng Viên Hướng Dẫn:", "ThS. Giảng Viên Hướng Dẫn (Khoa An toàn Thông tin - Đại học FPT)", "Email:", "truongnvse182034@fpt.edu.vn", "Không Gian:", "workspaces/truong_data_eng/")
    ]

    for row_idx, info in enumerate(student_info, start=5):
        ws1.row_dimensions[row_idx].height = 22
        if len(info) == 6:
            ws1.cell(row=row_idx, column=1, value=info[0]).font = bold_font
            ws1.cell(row=row_idx, column=2, value=info[1]).font = regular_font
            ws1.cell(row=row_idx, column=3, value=info[2]).font = bold_font
            ws1.cell(row=row_idx, column=4, value=info[3]).font = regular_font
            ws1.cell(row=row_idx, column=5, value=info[4]).font = bold_font
            ws1.cell(row=row_idx, column=6, value=info[5]).font = regular_font
        else:
            ws1.cell(row=row_idx, column=1, value=info[0]).font = bold_font
            ws1.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=7)
            ws1.cell(row=row_idx, column=2, value=info[1]).font = regular_font

    # Section 2: Nhiệm vụ chính
    ws1.merge_cells("A10:G10")
    ws1["A10"] = "II. TRỌNG TÂM NGHIÊN CỨU & KỸ THUẬT CÁ NHÂN"
    ws1["A10"].font = section_font
    ws1["A10"].fill = sub_fill
    ws1["A10"].alignment = left_align
    ws1.row_dimensions[10].height = 24

    responsibilities = [
        ("1. Điều phối & Quản trị Dự án:", "Giám sát tiến độ toàn nhóm, quản lý GitHub PR, đồng bộ các milestone và nộp báo cáo GVHD."),
        ("2. Kỹ thuật Dữ liệu (Data Engineering):", "Thu thập 5 tập dữ liệu Hugging Face, làm sạch, khử trùng lặp và xây dựng thuật toán Group-Aware Splitting chống rò rỉ dữ liệu."),
        ("3. Soạn thảo Luận văn Review 1:", "Chủ trì Chapter 1 (Introduction & Threat Model) và Chapter 2 (Literature Review & SOTA Survey)."),
        ("4. Kiểm thử & Đánh giá Đối chuẩn:", "Kiểm tra chéo các mô hình Baseline ML, DeBERTa-v3 Transformer và đo lường độ rò rỉ Inter-cluster Similarity.")
    ]

    for r_idx, (k, v) in enumerate(responsibilities, start=11):
        ws1.row_dimensions[r_idx].height = 22
        ws1.cell(row=r_idx, column=1, value=k).font = bold_font
        ws1.merge_cells(start_row=r_idx, start_column=2, end_row=r_idx, end_column=7)
        ws1.cell(row=r_idx, column=2, value=v).font = regular_font

    # Section 3: 4 Cột mốc cá nhân
    ws1.merge_cells("A16:G16")
    ws1["A16"] = "III. TIẾN ĐỘ 4 CỘT MỐC ĐÁNH GIÁ CỦA TRƯỞNG NHÓM"
    ws1["A16"].font = section_font
    ws1["A16"].fill = sub_fill
    ws1["A16"].alignment = left_align
    ws1.row_dimensions[16].height = 24

    milestone_headers = ["Cột Mốc Đánh Giá", "Thời Gian", "Nhiệm Vụ Cá Nhân Của Trường", "Sản Phẩm Đầu Ra", "Tỷ Trọng", "Trạng Thái", "Ghi Chú"]
    ws1.row_dimensions[17].height = 26
    for col_idx, h in enumerate(milestone_headers, start=1):
        c = ws1.cell(row=17, column=col_idx, value=h)
        c.font = header_font
        c.fill = primary_fill
        c.alignment = center_align
        c.border = thin_border

    milestones_data = [
        ("CỘT MỐC 1: REVIEW 1 (GVHD)", "Tuần 1 - 4 (07/09 - 04/10)", "Chủ trì Chapter 1, Chapter 2, Threat Model & Dàn ý Slide", "chapters/01 & 02, Hồ sơ Review 1", "35% Quá trình", "Đang thực hiện", "Hoàn thành Meeting 1 & 2, đang hoàn thiện 2 chương"),
        ("CỘT MỐC 2: REVIEW 2 (GVHD)", "Tuần 5 - 8 (05/10 - 01/11)", "Cào 5 datasets HF, Group-Aware Split, soạn thảo & cập nhật docs Chapter 3, bảo vệ Review 2 Tuần 8", "data/splits/, Report No.3 & docs Chapter 3", "20% Quá trình", "Chưa bắt đầu", "Bảo vệ Chapter 3 (Methodology) trước GVHD Tuần 8"),
        ("CỘT MỐC 3: HỘI ĐỒNG 1 (MIDTERM)", "Tuần 9 - 13 (02/11 - 06/12)", "Kiểm định rò rỉ dữ liệu, test tích hợp API Middleware & Demo", "Prototype hoàn chỉnh & Chapter 4", "Báo cáo Hội đồng 1 (Tuần 13)", "Chưa bắt đầu", "Bảo vệ Demo trước Hội đồng giữa kỳ Tuần 13"),
        ("CỘT MỐC 4: HỘI ĐỒNG FINAL", "Tuần 14 - 15 (07/12 - 20/12)", "Tổng hợp toàn văn 6 chương FINAL_THESIS & Slide bảo vệ", "FINAL_THESIS.md & Slide Final", "Bảo Vệ Tốt Nghiệp (Tuần 15)", "Chưa bắt đầu", "Bảo vệ chính thức trước Hội đồng Chấm thi Tuần 15")
    ]

    for r_idx, row in enumerate(milestones_data, start=18):
        ws1.row_dimensions[r_idx].height = 24
        for col_idx, val in enumerate(row, start=1):
            c = ws1.cell(row=r_idx, column=col_idx, value=val)
            c.font = regular_font
            c.border = thin_border
            if col_idx in [1, 2, 5, 6]:
                c.alignment = center_align
            else:
                c.alignment = left_align
            
            if val == "Hoàn thành":
                c.fill = green_fill
            elif val == "Đang thực hiện":
                c.fill = yellow_fill
            elif val == "Chưa bắt đầu":
                c.fill = gray_fill

    # ====================================================
    # SHEET 2: KẾ HOẠCH & NHẬT KÝ CÁ NHÂN
    # ====================================================
    ws2 = wb.create_sheet(title="2. Kế Hoạch & Nhật Ký")
    ws2.views.sheetView[0].showGridLines = True

    ws2.merge_cells("A1:G1")
    ws2["A1"] = "BẢNG PHÂN RÃ CÔNG VIỆC & NHẬT KÝ CHI TIẾT — NGUYỄN VĂN TRƯỜNG"
    ws2["A1"].font = title_font
    ws2["A1"].fill = primary_fill
    ws2["A1"].alignment = center_align
    ws2.row_dimensions[1].height = 32

    task_headers = ["Mã Task", "Tuần / Giai Đoạn", "Nội Dung Công Việc Chi Tiết Của Trường", "Trạng Thái", "Sản Phẩm Đầu Ra (Artifacts)", "Hạn Chót", "Đánh Giá Cá Nhân"]
    ws2.row_dimensions[3].height = 26
    for col_idx, h in enumerate(task_headers, start=1):
        c = ws2.cell(row=3, column=col_idx, value=h)
        c.font = header_font
        c.fill = primary_fill
        c.alignment = center_align
        c.border = thin_border

    # Data Validation Dropdown for Status
    dv = DataValidation(type="list", formula1='"Hoàn thành,Đang thực hiện,Chưa bắt đầu,Bị trễ"', allow_blank=True)
    ws2.add_data_validation(dv)
    dv.add("D4:D30")

    personal_tasks = [
        ("T01-NVT", "Tuần 1 (07/09 - 13/09)", "Họp khởi động GVHD (Meeting 1) & Họp nhóm sàng lọc 10 papers khoa học (Meeting 2)", "Hoàn thành", "Meeting 1 & Meeting 2.md", "13/09/2026", "Đã chốt mục tiêu và sàng lọc 10 papers"),
        ("T02-NVT", "Tuần 1 (07/09 - 13/09)", "Khảo sát và thu thập 17 bài báo IEEE/ACM >= 2022 về LLM Security", "Hoàn thành", "References/REFERENCES_LOG.md", "13/09/2026", "Đã lập ma trận liên kết 17 papers"),
        ("T03-NVT", "Tuần 2 (14/09 - 20/09)", "Soạn thảo Chapter 1: Background, Problem Statement (Von Neumann NLP)", "Đang thực hiện", "workspaces/truong_data_eng/docs/chapters/01", "16/09/2026", "Đang hoàn thiện phần bối cảnh"),
        ("T04-NVT", "Tuần 2 (14/09 - 20/09)", "Phân loại mối đe dọa (Threat Taxonomy: Direct/Indirect vs Jailbreak)", "Đang thực hiện", "workspaces/truong_data_eng/docs/thesis/", "17/09/2026", "Đã ánh xạ theo OWASP LLM01:2025"),
        ("T05-NVT", "Tuần 2 (14/09 - 20/09)", "Soạn thảo Chapter 2: Literature Review, SOTA Matrix & Research Gaps", "Đang thực hiện", "workspaces/truong_data_eng/docs/chapters/02", "19/09/2026", "Đã phân tích SOTA ProtectAI, NeMo"),
        ("T06-NVT", "Tuần 2 (14/09 - 20/09)", "Biên soạn Hồ sơ Kỹ thuật Review 1 & Dàn ý Slide 9 trang thuyết trình", "Đang thực hiện", "workspaces/truong_data_eng/docs/thesis/", "20/09/2026", "Chuẩn bị kịch bản thuyết trình 15 phút"),
        ("T07-NVT", "Tuần 3 (21/09 - 27/09)", "Điều phối hoàn thiện 2 chương, tổng kết và đồng bộ bản chính thức Review 1", "Đang thực hiện", "PI_GUARD_PROCESS_REPORT.xlsx", "27/09/2026", "Họp cả 4 thành viên thống nhất"),
        ("T08-NVT", "Tuần 4 (28/09 - 04/10)", "Chủ trì phần thuyết trình và BẢO VỆ REVIEW 1 TRƯỚC GVHD", "Chưa bắt đầu", "Biên bản nghiệm thu Review 1", "04/10/2026", "Bảo vệ thành công Chapter 1 & 2"),
        ("T09-NVT", "Tuần 5 - 6 (05/10 - 18/10)", "Tải 5 bộ dataset từ Hugging Face & viết thuật toán Group-Aware Split", "Chưa bắt đầu", "src/datasets/splitter.py", "18/10/2026", "Đảm bảo Jaccard similarity < 0.15"),
        ("T10-NVT", "Tuần 5 - 6 (05/10 - 18/10)", "Đánh giá phân phối nhãn, kiểm tra độ rò rỉ dữ liệu giữa Train/Val/Test", "Chưa bắt đầu", "notebooks/01_dataset_analysis", "18/10/2026", "Tránh rò rỉ mẫu tương tự"),
        ("T11-NVT", "Tuần 7 (19/10 - 25/10)", "Soạn thảo, cập nhật docs Chapter 3 (Methodology - Report No.3) & Slide Review 2", "Chưa bắt đầu", "docs/thesis/chapters/03_Methodology.md", "25/10/2026", "Chuẩn bị đầy đủ nội dung phương pháp luận"),
        ("T12-NVT", "Tuần 8 (26/10 - 01/11)", "BẢO VỆ REVIEW 2 TRƯỚC GVHD (CHAPTER 3), Nộp Report No.3 & Cập nhật Docs", "Chưa bắt đầu", "Report No.3 & Biên bản Review 2", "01/11/2026", "Bảo vệ Review 2 và cập nhật docs theo góp ý"),
        ("T13-NVT", "Tuần 9 - 12 (02/11 - 29/11)", "Hỗ trợ thử nghiệm nén Transformer INT8 & tích hợp Prototype", "Chưa bắt đầu", "models/onnx/ & Chapter 4", "29/11/2026", "Đo lường độ trễ trên CPU & chuẩn bị demo"),
        ("T14-NVT", "Tuần 13 (30/11 - 06/12)", "Điều phối kiểm thử tích hợp FastAPI Middleware & BẢO VỆ HỘI ĐỒNG 1", "Chưa bắt đầu", "Hệ thống Prototype Demo & Report No.4", "06/12/2026", "Bảo vệ Demo trước Hội đồng giữa kỳ Tuần 13"),
        ("T15-NVT", "Tuần 14 (07/12 - 13/12)", "Biên dịch toàn văn Master Thesis 6 chương qua script và quét Turnitin", "Chưa bắt đầu", "docs/thesis/FINAL_THESIS.md", "13/12/2026", "Đảm bảo độ trùng lặp Turnitin < 20%"),
        ("T16-NVT", "Tuần 15 (14/12 - 20/12)", "Chủ trì BẢO VỆ TỐT NGHIỆP CHÍNH THỨC TRƯỚC HỘI ĐỒNG FPT", "Chưa bắt đầu", "Slide Bảo vệ & Đồ án hoàn chỉnh", "20/12/2026", "Tốt nghiệp đạt điểm xuất sắc")
    ]

    for r_idx, task in enumerate(personal_tasks, start=4):
        ws2.row_dimensions[r_idx].height = 22
        for col_idx, val in enumerate(task, start=1):
            c = ws2.cell(row=r_idx, column=col_idx, value=val)
            c.font = regular_font
            c.border = thin_border
            if col_idx in [1, 2, 4, 6]:
                c.alignment = center_align
            else:
                c.alignment = left_align
            
            if col_idx == 4:
                if val == "Hoàn thành":
                    c.fill = green_fill
                    c.font = bold_font
                elif val == "Đang thực hiện":
                    c.fill = yellow_fill
                    c.font = bold_font
                elif val == "Chưa bắt đầu":
                    c.fill = gray_fill

    # Auto-fit Column Widths across all sheets
    ws1.column_dimensions['A'].width = 24
    ws1.column_dimensions['B'].width = 30
    ws1.column_dimensions['C'].width = 20
    ws1.column_dimensions['D'].width = 28
    ws1.column_dimensions['E'].width = 18
    ws1.column_dimensions['F'].width = 20
    ws1.column_dimensions['G'].width = 38

    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 24
    ws2.column_dimensions['C'].width = 48
    ws2.column_dimensions['D'].width = 18
    ws2.column_dimensions['E'].width = 38
    ws2.column_dimensions['F'].width = 14
    ws2.column_dimensions['G'].width = 34

    wb.save(output_path)
    print(f"Personal Process Report saved to: {output_path}")

if __name__ == "__main__":
    create_personal_process_report()
