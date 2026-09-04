"""
PI-Guard Google Sheets Synchronization Utility
==============================================
Tự động đồng bộ tiến độ giữa Google Sheets trực tuyến và các file cục bộ trong repo:
- Google Sheet: https://docs.google.com/spreadsheets/d/1toBlihg3ycJq_0uOLbfHdD3NPMLCYWt1Az460sAfl2M/edit?gid=758037622
- Local Excel: Meeting/PI_GUARD_PROCESS_REPORT.xlsx & reports/PI_GUARD_PROCESS_REPORT.xlsx
"""

import sys
import os
import csv
import io
import urllib.request
import openpyxl
from openpyxl.styles import PatternFill, Font

if sys.stdout.encoding != 'utf-8':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

SHEET_ID = "1toBlihg3ycJq_0uOLbfHdD3NPMLCYWt1Az460sAfl2M"
GID = "758037622"
EXPORT_CSV_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=csv&gid={GID}"

WORKSPACE_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATHS = [
    os.path.join(WORKSPACE_ROOT, "reports", "PI_GUARD_PROCESS_REPORT.xlsx"),
    os.path.join(WORKSPACE_ROOT, "Meeting", "PI_GUARD_PROCESS_REPORT.xlsx")
]

def fetch_online_tasks():
    """Tải dữ liệu mới nhất từ Google Sheet qua CSV export endpoint"""
    print(f"[*] Đang tải dữ liệu từ Google Sheets: {EXPORT_CSV_URL} ...")
    req = urllib.request.Request(EXPORT_CSV_URL, headers={'User-Agent': 'Mozilla/5.0'})
    with urllib.request.urlopen(req) as resp:
        content = resp.read().decode('utf-8')
    
    reader = csv.reader(io.StringIO(content))
    rows = list(reader)
    
    tasks = []
    header_found = False
    for row in rows:
        if not row:
            continue
        if len(row) >= 4 and row[0].strip() == "Mã Task":
            header_found = True
            continue
        if header_found and len(row) >= 7 and row[0].strip().startswith("T"):
            tasks.append({
                "id": row[0].strip(),
                "phase": row[1].strip(),
                "title": row[2].strip(),
                "status": row[3].strip(),
                "assignee": row[4].strip(),
                "deliverable": row[5].strip(),
                "deadline": row[6].strip()
            })
    
    print(f"[+] Đã tải thành công {len(tasks)} đầu việc từ Google Sheets!")
    return tasks

def update_local_excel(tasks):
    """Cập nhật trạng thái từ Google Sheets vào file Excel cục bộ"""
    for excel_path in EXCEL_PATHS:
        if not os.path.exists(excel_path):
            print(f"[!] Không tìm thấy file {excel_path}, bỏ qua.")
            continue
        
        try:
            wb = openpyxl.load_workbook(excel_path)
            if '2. Checklist Tiến Độ' in wb.sheetnames:
                ws = wb['2. Checklist Tiến Độ']
                for r in range(5, ws.max_row + 1):
                    task_id_cell = ws.cell(row=r, column=1).value
                    if not task_id_cell:
                        continue
                    matched = next((t for t in tasks if t['id'] == str(task_id_cell).strip()), None)
                    if matched:
                        status_cell = ws.cell(row=r, column=4)
                        status_cell.value = matched['status']
                        if matched['status'] == 'Hoàn thành':
                            status_cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
                            status_cell.font = Font(name='Calibri', size=10, bold=True, color='375623')
                        elif matched['status'] == 'Đang thực hiện':
                            status_cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
                            status_cell.font = Font(name='Calibri', size=10, bold=True, color='7F6000')
                        else:
                            status_cell.fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')
                            status_cell.font = Font(name='Calibri', size=10)
                
                wb.save(excel_path)
                print(f"[+] Đã cập nhật file Excel: {excel_path}")
        except Exception as e:
            print(f"[!] Lỗi khi cập nhật file Excel {excel_path}: {e}")

def print_summary(tasks):
    """In bảng tổng kết trạng thái công việc ra console"""
    completed = [t for t in tasks if t['status'] == 'Hoàn thành']
    in_progress = [t for t in tasks if t['status'] == 'Đang thực hiện']
    pending = [t for t in tasks if t['status'] == 'Chưa bắt đầu']
    
    print("\n" + "="*70)
    print("📊 BẢNG TỔNG KẾT TIẾN ĐỘ GOOGLE SHEETS (PI-GUARD CAPSTONE)")
    print("="*70)
    print(f"Tổng số Tasks: {len(tasks)}")
    print(f"✅ Đã hoàn thành: {len(completed)} task(s)")
    print(f"🔄 Đang thực hiện: {len(in_progress)} task(s)")
    print(f"⏳ Chưa bắt đầu:  {len(pending)} task(s)")
    print("-"*70)
    print(f"{'MÃ':<6} | {'TRẠNG THÁI':<16} | {'NGƯỜI PHỤ TRÁCH':<18} | {'NỘI DUNG'}")
    print("-"*70)
    for t in tasks:
        print(f"{t['id']:<6} | {t['status']:<16} | {t['assignee']:<18} | {t['title'][:35]}...")
    print("="*70 + "\n")

if __name__ == "__main__":
    tasks = fetch_online_tasks()
    if tasks:
        print_summary(tasks)
        update_local_excel(tasks)
